import csv
import os
from datetime import datetime

from openpyxl import Workbook, load_workbook


def safe_print(message):
    try:
        print(message)
    except UnicodeEncodeError:
        print(str(message).encode("ascii", "ignore").decode("ascii"))


class Logger:
    def __init__(self, file_path="data/logs/applications.csv", excel_path="data/logs/applications.xlsx"):
        self.file_path = file_path
        self.excel_path = excel_path
        self.headers = ["Timestamp", "Job Title", "Company", "URL", "Score", "Status", "Reason"]
        self._initialize_log()

    def _initialize_log(self):
        os.makedirs(os.path.dirname(self.file_path), exist_ok=True)

        if not os.path.exists(self.file_path):
            with open(self.file_path, mode="w", newline="", encoding="utf-8") as handle:
                writer = csv.writer(handle)
                writer.writerow(self.headers)

        if not os.path.exists(self.excel_path):
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Applications"
            sheet.append(self.headers)
            workbook.save(self.excel_path)

    def _append_excel_row(self, row):
        workbook = load_workbook(self.excel_path)
        sheet = workbook.active
        sheet.append(row)
        workbook.save(self.excel_path)

    def log_application(self, job_title, company, url, score, status, reason=""):
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        row = [timestamp, job_title, company, url, score, status, reason]

        try:
            with open(self.file_path, mode="a", newline="", encoding="utf-8") as handle:
                writer = csv.writer(handle)
                writer.writerow(row)
            self._append_excel_row(row)
        except Exception as exc:
            safe_print(f"File logging error: {exc}")

        safe_print(f"Logged: {job_title} at {company} - {status}")
        safe_print(f"Saved tracker row to: {self.excel_path}")
